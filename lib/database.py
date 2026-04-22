"""
SQLite operations for the European Custom Database and the simulation DB.

Schema is shared; the simulation DB adds a `fired` column to track
which March-2026 transactions have been replayed.
"""
from __future__ import annotations

import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from lib.config import EUROPEAN_CUSTOM_DB, SIMULATION_DB, INVESTIGATION_DB, SEED_CASES_DB, HISTORICAL_CASES_DB

# ── Schema ────────────────────────────────────────────────────────────────────

_TX_DDL = """
CREATE TABLE IF NOT EXISTS transactions (
    id               INTEGER PRIMARY KEY AUTOINCREMENT,
    transaction_id   TEXT    UNIQUE NOT NULL,
    transaction_date TEXT    NOT NULL,
    seller_id        TEXT    NOT NULL,
    seller_name      TEXT    NOT NULL,
    seller_country   TEXT    NOT NULL,
    item_description TEXT    NOT NULL,
    item_category    TEXT    NOT NULL,
    value            REAL    NOT NULL,
    vat_rate         REAL    NOT NULL,
    vat_amount       REAL    NOT NULL,
    buyer_country    TEXT    NOT NULL,
    correct_vat_rate REAL    NOT NULL,
    has_error        INTEGER NOT NULL DEFAULT 0,
    xml_message      TEXT,
    created_at       TEXT    NOT NULL,
    -- Producer (non-EU manufacturer) sourced by the seller/reseller.
    -- Populated by the seeder for new rows. May be NULL on rows from
    -- older DBs created before the two-tier party model was introduced.
    producer_id      TEXT,
    producer_name    TEXT,
    producer_country TEXT,
    producer_city    TEXT
);
CREATE INDEX IF NOT EXISTS idx_tx_date     ON transactions(transaction_date);
CREATE INDEX IF NOT EXISTS idx_tx_seller   ON transactions(seller_name);
CREATE INDEX IF NOT EXISTS idx_tx_buyer    ON transactions(buyer_country);
CREATE INDEX IF NOT EXISTS idx_tx_producer ON transactions(producer_country);
"""

_SIM_DDL = _TX_DDL + """
ALTER TABLE transactions ADD COLUMN fired INTEGER NOT NULL DEFAULT 0;
CREATE INDEX IF NOT EXISTS idx_fired ON transactions(fired, transaction_date);
"""

_ALARM_DDL = """
CREATE TABLE IF NOT EXISTS alarms (
    id               INTEGER PRIMARY KEY AUTOINCREMENT,
    alarm_key        TEXT    NOT NULL,
    supplier_id      TEXT    NOT NULL,
    supplier_name    TEXT    NOT NULL,
    buyer_country    TEXT    NOT NULL,
    trigger_tx_id    TEXT    NOT NULL,
    raised_at        TEXT    NOT NULL,
    expires_at       TEXT    NOT NULL,
    ratio_current    REAL    NOT NULL,
    ratio_historical REAL    NOT NULL,
    deviation_pct    REAL    NOT NULL,
    active           INTEGER NOT NULL DEFAULT 1
);
CREATE INDEX IF NOT EXISTS idx_alarm_key     ON alarms(alarm_key, active);
CREATE INDEX IF NOT EXISTS idx_alarm_expires ON alarms(expires_at);
"""

_AGENT_LOG_DDL = """
CREATE TABLE IF NOT EXISTS agent_log (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    transaction_id    TEXT,
    seller_name       TEXT,
    buyer_country     TEXT,
    item_description  TEXT,
    item_category     TEXT,
    value             REAL,
    vat_rate          REAL,
    correct_vat_rate  REAL,
    verdict           TEXT,
    reasoning         TEXT,
    legislation_refs  TEXT,
    sent_to_ireland   INTEGER DEFAULT 0,
    processed_at      TEXT
);
CREATE INDEX IF NOT EXISTS idx_agent_log_tx ON agent_log(transaction_id);
"""

_IRELAND_QUEUE_DDL = """
CREATE TABLE IF NOT EXISTS ireland_queue (
    id               INTEGER PRIMARY KEY AUTOINCREMENT,
    transaction_id   TEXT    NOT NULL,
    seller_name      TEXT    NOT NULL,
    seller_country   TEXT    NOT NULL,
    item_description TEXT    NOT NULL,
    item_category    TEXT    NOT NULL,
    value            REAL    NOT NULL,
    vat_rate         REAL    NOT NULL,
    correct_vat_rate REAL    NOT NULL,
    vat_amount       REAL    NOT NULL,
    transaction_date TEXT    NOT NULL,
    alarm_key        TEXT    NOT NULL,
    deviation_pct    REAL,
    ratio_current    REAL,
    ratio_historical REAL,
    agent_verdict    TEXT    NOT NULL,
    agent_reasoning  TEXT    NOT NULL,
    queued_at        TEXT    NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_ireland_queue_tx ON ireland_queue(transaction_id);
"""


# ── Reference tables (lookups previously hardcoded in the frontend) ──────────
#
# Hosted in european_custom.db. Read by GET /api/reference; seeded on first
# init from the constants below (idempotent — INSERT OR IGNORE).

_VAT_CATEGORIES_DDL = """
CREATE TABLE IF NOT EXISTS vat_categories (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    label       TEXT    UNIQUE NOT NULL,
    rate        REAL    NOT NULL,
    description TEXT,
    sort_order  INTEGER NOT NULL DEFAULT 0
);
"""

_RISK_LEVELS_DDL = """
CREATE TABLE IF NOT EXISTS risk_levels (
    name          TEXT    PRIMARY KEY,
    display_color TEXT,
    sort_order    INTEGER NOT NULL DEFAULT 0
);
"""

_EU_REGIONS_DDL = """
CREATE TABLE IF NOT EXISTS eu_regions (
    country_code TEXT    PRIMARY KEY,
    country_name TEXT,
    region       TEXT    NOT NULL
);
"""

_SUSPICION_TYPES_DDL = """
CREATE TABLE IF NOT EXISTS suspicion_types (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    name        TEXT    UNIQUE NOT NULL,
    description TEXT,
    icon        TEXT,
    color       TEXT,
    sort_order  INTEGER NOT NULL DEFAULT 0
);
"""

_SALES_ORDER_STATUSES_DDL = """
CREATE TABLE IF NOT EXISTS sales_order_statuses (
    name        TEXT    PRIMARY KEY,
    description TEXT,
    sort_order  INTEGER NOT NULL DEFAULT 0
);
"""

_SEED_SALES_ORDER_STATUSES = [
    ("Under Investigation", "Set at record creation when the C&T factory opens a case", 10),
    ("To Be Released",      "Set by the Customs officer when recommending release",     20),
    ("To Be Retained",      "Set by the Customs officer when recommending retainment",  30),
]


_CASE_STATUSES_DDL = """
CREATE TABLE IF NOT EXISTS case_statuses (
    name        TEXT    PRIMARY KEY,
    description TEXT,
    sort_order  INTEGER NOT NULL DEFAULT 0
);
"""

_SEED_CASE_STATUSES = [
    ("New",                            "Set at case creation by the C&T factory",                       10),
    ("Under Review by Customs",        "Customs officer has opened / is reviewing the case",            20),
    ("AI Investigation in Progress",   "VAT Fraud Detection agent is analysing (Tax review requested)", 30),
    ("Under Review by Tax",            "Tax officer is reviewing after AI analysis completes",          40),
    ("Requested Input by Third Party", "Awaiting response from an external third party",                50),
    ("Closed",                         "Terminal — officer recommended release or retainment",          60),
]


_SEED_VAT_CATEGORIES = [
    ("Educational Material", 9.0,  "Books, learning aids", 10),
    ("Consumer Electronics", 23.0, "Phones, audio, smart devices", 20),
    ("Fashion & Apparel",    23.0, "Clothing, footwear, accessories", 30),
    ("Health & Beauty",      13.5, "Cosmetics, personal care", 40),
    ("Home & Garden",        23.0, "Appliances, furniture, decor", 50),
    ("Accessories",          23.0, "Phone/computer accessories", 60),
    ("Toys & Games",         13.5, "Toys, board games", 70),
]

_SEED_RISK_LEVELS = [
    ("Critical", "destructive", 10),
    ("High",     "risk-high",   20),
    ("Medium",   "warning",     30),
    ("Low",      "success",     40),
]

_SEED_REGIONS = [
    # Ireland
    ("IE", "Ireland",         "Ireland"),
    ("GB", "United Kingdom",  "Ireland"),
    # Western EU
    ("FR", "France",          "Western EU"),
    ("BE", "Belgium",         "Western EU"),
    ("NL", "Netherlands",     "Western EU"),
    ("LU", "Luxembourg",      "Western EU"),
    ("DE", "Germany",         "Western EU"),
    # Southern EU
    ("ES", "Spain",           "Southern EU"),
    ("PT", "Portugal",        "Southern EU"),
    ("IT", "Italy",           "Southern EU"),
    # Central / Eastern EU
    ("PL", "Poland",          "Central EU"),
    ("CZ", "Czech Republic",  "Central EU"),
    ("HU", "Hungary",         "Central EU"),
    ("SK", "Slovakia",        "Central EU"),
    # Nordics
    ("DK", "Denmark",         "Nordics"),
    ("SE", "Sweden",          "Nordics"),
    ("FI", "Finland",         "Nordics"),
]

_ML_RISK_RULES_DDL = """
CREATE TABLE IF NOT EXISTS ml_risk_rules (
    id                           INTEGER PRIMARY KEY AUTOINCREMENT,
    seller                       TEXT NOT NULL,
    country_origin               TEXT NOT NULL,
    vat_product_category         TEXT NOT NULL,
    country_destination          TEXT NOT NULL,
    risk                         REAL NOT NULL,
    description                  TEXT,
    seller_weight                REAL,
    country_origin_weight        REAL,
    vat_product_category_weight  REAL,
    country_destination_weight   REAL,
    UNIQUE(seller, country_origin, vat_product_category, country_destination)
);
CREATE INDEX IF NOT EXISTS idx_mlrr_lookup
    ON ml_risk_rules(seller, country_origin, vat_product_category, country_destination);
"""


_SEED_SUSPICION_TYPES = [
    ("VAT Rate Deviation",
     "Goods reported at differing VAT rates across shipments — pointing to rate misclassification or selective underreporting.",
     "AlertTriangle", "risk-critical", 10),
    ("Customs Duty Gap",
     "Customs duties declared differ from the expected tariff for the declared commodity code.",
     "FileWarning",   "risk-high",     20),
    ("Product Type Mismatch",
     "Commodity description conflicts with the product category used in the IOSS VAT filing.",
     "Package",       "risk-medium",   30),
    ("Taxable Value Understatement",
     "Declared item value appears lower than market value for the product category.",
     "DollarSign",    "risk-high",     40),
    ("Watchlist Match",
     "Seller, supplier, or origin matches an active enforcement watchlist.",
     "ShieldAlert",   "risk-critical", 50),
]


# ── Data hub schema (3 dark-purple tables from the data model diagram) ───────
#
# ── New data model (Entity Data Model Simplified) ─────────────────────────────
# Three tables with 1:1 relationships keyed on Sales_Order_Business_Key.
# Sales_Order and Sales_Order_Risk live in european_custom.db (data hub).
# Sales_Order_Case lives in a separate investigation.db.
# Field names match the JSON event payloads and the data model diagram.

_SALES_ORDER_DDL = """
CREATE TABLE IF NOT EXISTS Sales_Order (
    Sales_Order_ID              TEXT NOT NULL,
    Sales_Order_Business_Key    TEXT PRIMARY KEY,
    HS_Product_Category         TEXT,
    Product_Description         TEXT,
    Product_Value               REAL,
    VAT_Rate                    REAL,
    VAT_Fee                     REAL,
    Seller_Name                 TEXT,
    Country_Origin              TEXT,
    Country_Destination         TEXT,
    Status                      TEXT,
    Update_time                 TEXT,
    Updated_by                  TEXT,
    Case_ID                     TEXT
);
CREATE INDEX IF NOT EXISTS idx_so_id      ON Sales_Order(Sales_Order_ID);
CREATE INDEX IF NOT EXISTS idx_so_status  ON Sales_Order(Status);
CREATE INDEX IF NOT EXISTS idx_so_origin  ON Sales_Order(Country_Origin);
CREATE INDEX IF NOT EXISTS idx_so_dest    ON Sales_Order(Country_Destination);
CREATE INDEX IF NOT EXISTS idx_so_case_id ON Sales_Order(Case_ID);
"""

_SALES_ORDER_RISK_DDL = """
CREATE TABLE IF NOT EXISTS Sales_Order_Risk (
    Sales_Order_Risk_ID         TEXT PRIMARY KEY,
    Sales_Order_Business_Key    TEXT NOT NULL,
    Risk_Type                   TEXT,
    Overall_Risk_Score          REAL,
    Overall_Risk_Level          TEXT,
    Seller_Risk_Score           REAL,
    Country_Risk_Score          REAL,
    Product_Category_Risk_Score REAL,
    Manufacturer_Risk_Score     REAL,
    Confidence_Score            REAL,
    Overall_Risk_Description    TEXT,
    Proposed_Risk_Action        TEXT,
    Risk_Comment                TEXT,
    Evaluation_by               TEXT,
    Update_time                 TEXT,
    Updated_by                  TEXT,
    FOREIGN KEY (Sales_Order_Business_Key)
        REFERENCES Sales_Order(Sales_Order_Business_Key)
);
CREATE INDEX IF NOT EXISTS idx_sor_bk    ON Sales_Order_Risk(Sales_Order_Business_Key);
CREATE INDEX IF NOT EXISTS idx_sor_level ON Sales_Order_Risk(Overall_Risk_Level);
"""

_SALES_ORDER_CASE_DDL = """
CREATE TABLE IF NOT EXISTS Sales_Order_Case (
    Case_ID                          TEXT PRIMARY KEY,
    Sales_Order_Business_Key         TEXT NOT NULL,
    Status                           TEXT,
    VAT_Problem_Type                 TEXT,
    Recommended_Product_Value        REAL,
    Recommended_VAT_Product_Category TEXT,
    Recommended_VAT_Rate             REAL,
    Recommended_VAT_Fee              REAL,
    AI_Analysis                      TEXT,
    AI_Confidence                    REAL,
    VAT_Gap_Fee                      REAL,
    Evaluation_by                    TEXT,
    Proposed_Action_Tax              TEXT,
    Proposed_Action_Customs          TEXT,
    Communication                    TEXT,
    Additional_Evidence              TEXT,
    Update_time                      TEXT,
    Updated_by                       TEXT,
    Created_time                     TEXT,
    -- Case-level overall risk score (0-1, averaged across all orders)
    Overall_Case_Risk_Score          REAL DEFAULT 0,
    Overall_Case_Risk_Level          TEXT DEFAULT 'Low',
    -- Per-engine risk scores (0-1 average across all orders in the case)
    Engine_VAT_Ratio                 REAL DEFAULT 0,
    Engine_ML_Watchlist              REAL DEFAULT 0,
    Engine_IE_Seller_Watchlist       REAL DEFAULT 0,
    Engine_Description_Vagueness     REAL DEFAULT 0
);
CREATE INDEX IF NOT EXISTS idx_soc_bk      ON Sales_Order_Case(Sales_Order_Business_Key);
CREATE INDEX IF NOT EXISTS idx_soc_status  ON Sales_Order_Case(Status);
CREATE INDEX IF NOT EXISTS idx_soc_created ON Sales_Order_Case(Created_time);

CREATE TABLE IF NOT EXISTS risk_engine_signals (
    field_name    TEXT PRIMARY KEY,
    engine_key    TEXT NOT NULL,
    display_name  TEXT NOT NULL,
    description   TEXT
);
INSERT OR IGNORE INTO risk_engine_signals VALUES
    ('Engine_VAT_Ratio',             'vat_ratio',              'VAT Ratio Deviation',        'Statistical deviation in VAT/value ratio vs 8-week baseline'),
    ('Engine_ML_Watchlist',          'watchlist',              'VAT Misclassification Risk',  'ML-based watchlist matching on seller × origin × category × destination'),
    ('Engine_IE_Seller_Watchlist',   'ireland_watchlist',      'Seller Risk',                 'Ireland-specific seller watchlist for IE-destined goods'),
    ('Engine_Description_Vagueness', 'description_vagueness',  'Description Vagueness',       'NLP-based detection of vague or generic product descriptions');
"""



def _connect(path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(path, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    return conn


_NEW_DATASET_TX_COLUMNS: list[tuple[str, str]] = [
    # New-dataset (Stage 3) per-tx fields. All nullable so the legacy
    # seeder still works — when these are NULL, the engines fall back
    # to their legacy paths (volume-ratio alarm, 4-tuple ML lookup,
    # embedding-based vagueness).
    ("vat_subcategory_code",              "TEXT  DEFAULT NULL"),
    ("engine_vat_ratio_risk",             "REAL  DEFAULT NULL"),
    ("engine_ml_risk",                    "REAL  DEFAULT NULL"),
    ("engine_ml_seller_contribution",     "REAL  DEFAULT NULL"),
    ("engine_ml_origin_contribution",     "REAL  DEFAULT NULL"),
    ("engine_ml_category_contribution",   "REAL  DEFAULT NULL"),
    ("engine_ml_destination_contribution","REAL  DEFAULT NULL"),
    ("engine_vagueness_risk",             "REAL  DEFAULT NULL"),
    ("engine_ie_watchlist_risk",          "REAL  DEFAULT NULL"),
]


def _migrate_european_custom_db(conn: sqlite3.Connection) -> None:
    """Add columns / tables introduced after initial schema."""
    for col, definition in [
        ("suspicious",       "INTEGER DEFAULT 0"),
        ("alarm_id",         "INTEGER DEFAULT NULL"),
        ("suspicion_level",  "TEXT    DEFAULT NULL"),
        # Two-tier party model — non-EU producer (the line item Seller).
        ("producer_id",      "TEXT    DEFAULT NULL"),
        ("producer_name",    "TEXT    DEFAULT NULL"),
        ("producer_country", "TEXT    DEFAULT NULL"),
        ("producer_city",    "TEXT    DEFAULT NULL"),
        *_NEW_DATASET_TX_COLUMNS,
    ]:
        try:
            conn.execute(f"ALTER TABLE transactions ADD COLUMN {col} {definition}")
        except sqlite3.OperationalError:
            pass   # already exists
    for ddl in [
        _ALARM_DDL, _AGENT_LOG_DDL, _IRELAND_QUEUE_DDL,
    ]:
        for stmt in ddl.strip().split(";"):
            s = stmt.strip()
            if s:
                conn.execute(s)


def _migrate_simulation_db(conn: sqlite3.Connection) -> None:
    """Add producer + new-dataset columns to existing simulation.db files
    (the production table column list lives in _TX_DDL but older DB files
    predate them)."""
    for col, definition in [
        ("producer_id",      "TEXT DEFAULT NULL"),
        ("producer_name",    "TEXT DEFAULT NULL"),
        ("producer_country", "TEXT DEFAULT NULL"),
        ("producer_city",    "TEXT DEFAULT NULL"),
        *_NEW_DATASET_TX_COLUMNS,
    ]:
        try:
            conn.execute(f"ALTER TABLE transactions ADD COLUMN {col} {definition}")
        except sqlite3.OperationalError:
            pass   # already exists


def init_european_custom_db() -> None:
    conn = _connect(EUROPEAN_CUSTOM_DB)
    with conn:
        # Run the CREATE TABLE first (without indexes that reference
        # columns added by migration). Then migrate to add any missing
        # columns. Finally create the indexes.
        for stmt in _TX_DDL.strip().split(";"):
            s = stmt.strip()
            if not s:
                continue
            # Skip CREATE INDEX statements on first pass — columns
            # they reference may not exist yet on older DBs.
            if s.upper().startswith("CREATE INDEX"):
                continue
            try:
                conn.execute(s)
            except sqlite3.OperationalError:
                pass  # table already exists
        # Add any missing columns (producer_*, suspicious, etc.)
        _migrate_european_custom_db(conn)
        # Now create indexes — all columns are guaranteed to exist.
        for stmt in _TX_DDL.strip().split(";"):
            s = stmt.strip()
            if s and s.upper().startswith("CREATE INDEX"):
                try:
                    conn.execute(s)
                except sqlite3.OperationalError:
                    pass  # index already exists
    conn.close()
    # Backfill the new data hub table from the legacy transactions table.
    # Idempotent (uses INSERT OR REPLACE keyed on the synthetic SKU), so it's
    # Create the new data model tables (Sales_Order + Sales_Order_Risk)
    _init_ddl(EUROPEAN_CUSTOM_DB, _SALES_ORDER_DDL)
    _init_ddl(EUROPEAN_CUSTOM_DB, _SALES_ORDER_RISK_DDL)
    # Reference / lookup tables (replace static frontend constants)
    _init_ddl(EUROPEAN_CUSTOM_DB, _VAT_CATEGORIES_DDL)
    _init_ddl(EUROPEAN_CUSTOM_DB, _RISK_LEVELS_DDL)
    _init_ddl(EUROPEAN_CUSTOM_DB, _EU_REGIONS_DDL)
    _init_ddl(EUROPEAN_CUSTOM_DB, _SUSPICION_TYPES_DDL)
    _init_ddl(EUROPEAN_CUSTOM_DB, _ML_RISK_RULES_DDL)
    _init_ddl(EUROPEAN_CUSTOM_DB, _SALES_ORDER_STATUSES_DDL)
    _init_ddl(EUROPEAN_CUSTOM_DB, _CASE_STATUSES_DDL)
    _seed_reference_tables()
    _seed_ml_risk_rules_from_xlsx()


def init_historical_cases_db() -> None:
    """Create the 3-table case dataset in historical_cases.db — same
    shape as investigation.db. Holds PAST closed cases (IE destination)
    used as the source for /api/rg/cases/{id}/previous."""
    _init_ddl(HISTORICAL_CASES_DB, _SALES_ORDER_DDL)
    _init_ddl(HISTORICAL_CASES_DB, _SALES_ORDER_RISK_DDL)
    _init_ddl(HISTORICAL_CASES_DB, _SALES_ORDER_CASE_DDL)


def init_investigation_db() -> None:
    """Create the 3-table case dataset (Sales_Order + Sales_Order_Risk +
    Sales_Order_Case) in investigation.db. These mirror the data model used
    by the data hub and give the C&T Risk Management System a self-contained store."""
    _init_ddl(INVESTIGATION_DB, _SALES_ORDER_DDL)
    _init_ddl(INVESTIGATION_DB, _SALES_ORDER_RISK_DDL)
    _init_ddl(INVESTIGATION_DB, _SALES_ORDER_CASE_DDL)
    # Migrate older DBs: add columns introduced after initial schema
    conn = _connect(INVESTIGATION_DB)
    with conn:
        for table, col, definition in [
            ("Sales_Order_Case", "Created_time",              "TEXT"),
            ("Sales_Order",      "Case_ID",                   "TEXT"),
            ("Sales_Order_Case", "Overall_Case_Risk_Score",   "REAL DEFAULT 0"),
            ("Sales_Order_Case", "Overall_Case_Risk_Level",   "TEXT DEFAULT 'Low'"),
            ("Sales_Order_Case", "Engine_VAT_Ratio",          "REAL DEFAULT 0"),
            ("Sales_Order_Case", "Engine_ML_Watchlist",       "REAL DEFAULT 0"),
            ("Sales_Order_Case", "Engine_IE_Seller_Watchlist", "REAL DEFAULT 0"),
            ("Sales_Order_Case", "Engine_Description_Vagueness", "REAL DEFAULT 0"),
        ]:
            try:
                conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} {definition}")
            except sqlite3.OperationalError:
                pass
        conn.execute("UPDATE Sales_Order_Case SET Created_time = Update_time WHERE Created_time IS NULL")
    conn.close()


def _init_ddl(db_path, ddl: str) -> None:
    """Two-pass DDL init: tables first, then indexes."""
    conn = _connect(db_path)
    with conn:
        for stmt in ddl.strip().split(";"):
            s = stmt.strip()
            if not s or s.upper().startswith("CREATE INDEX"):
                continue
            try:
                conn.execute(s)
            except sqlite3.OperationalError:
                pass
        for stmt in ddl.strip().split(";"):
            s = stmt.strip()
            if s and s.upper().startswith("CREATE INDEX"):
                try:
                    conn.execute(s)
                except sqlite3.OperationalError:
                    pass
    conn.close()


def upsert_sales_order(row: dict) -> None:
    """Insert or update a Sales_Order record."""
    conn = _connect(EUROPEAN_CUSTOM_DB)
    with conn:
        conn.execute("""
            INSERT OR REPLACE INTO Sales_Order (
                Sales_Order_ID, Sales_Order_Business_Key,
                HS_Product_Category, Product_Description, Product_Value,
                VAT_Rate, VAT_Fee, Seller_Name,
                Country_Origin, Country_Destination,
                Status, Update_time, Updated_by
            ) VALUES (
                :Sales_Order_ID, :Sales_Order_Business_Key,
                :HS_Product_Category, :Product_Description, :Product_Value,
                :VAT_Rate, :VAT_Fee, :Seller_Name,
                :Country_Origin, :Country_Destination,
                :Status, :Update_time, :Updated_by
            )
        """, row)
    conn.close()


def upsert_sales_order_risk(row: dict) -> None:
    """Insert or update a Sales_Order_Risk record."""
    conn = _connect(EUROPEAN_CUSTOM_DB)
    with conn:
        conn.execute("""
            INSERT OR REPLACE INTO Sales_Order_Risk (
                Sales_Order_Risk_ID, Sales_Order_Business_Key,
                Risk_Type, Overall_Risk_Score, Overall_Risk_Level,
                Seller_Risk_Score, Country_Risk_Score,
                Product_Category_Risk_Score, Manufacturer_Risk_Score,
                Confidence_Score, Overall_Risk_Description,
                Proposed_Risk_Action, Risk_Comment,
                Evaluation_by, Update_time, Updated_by
            ) VALUES (
                :Sales_Order_Risk_ID, :Sales_Order_Business_Key,
                :Risk_Type, :Overall_Risk_Score, :Overall_Risk_Level,
                :Seller_Risk_Score, :Country_Risk_Score,
                :Product_Category_Risk_Score, :Manufacturer_Risk_Score,
                :Confidence_Score, :Overall_Risk_Description,
                :Proposed_Risk_Action, :Risk_Comment,
                :Evaluation_by, :Update_time, :Updated_by
            )
        """, row)
    conn.close()


def upsert_sales_order_case(row: dict) -> None:
    """Insert or update a Sales_Order_Case record in the investigation DB."""
    conn = _connect(INVESTIGATION_DB)
    with conn:
        conn.execute("""
            INSERT OR REPLACE INTO Sales_Order_Case (
                Case_ID, Sales_Order_Business_Key, Status,
                VAT_Problem_Type, Recommended_Product_Value,
                Recommended_VAT_Product_Category, Recommended_VAT_Rate,
                Recommended_VAT_Fee, AI_Analysis, AI_Confidence,
                VAT_Gap_Fee, Evaluation_by,
                Proposed_Action_Tax, Proposed_Action_Customs,
                Communication, Additional_Evidence,
                Update_time, Updated_by
            ) VALUES (
                :Case_ID, :Sales_Order_Business_Key, :Status,
                :VAT_Problem_Type, :Recommended_Product_Value,
                :Recommended_VAT_Product_Category, :Recommended_VAT_Rate,
                :Recommended_VAT_Fee, :AI_Analysis, :AI_Confidence,
                :VAT_Gap_Fee, :Evaluation_by,
                :Proposed_Action_Tax, :Proposed_Action_Customs,
                :Communication, :Additional_Evidence,
                :Update_time, :Updated_by
            )
        """, row)
    conn.close()


def init_simulation_db() -> None:
    conn = _connect(SIMULATION_DB)
    with conn:
        # Two-pass DDL: tables first, then migrate, then indexes.
        for stmt in _TX_DDL.strip().split(";"):
            s = stmt.strip()
            if not s or s.upper().startswith("CREATE INDEX"):
                continue
            try:
                conn.execute(s)
            except sqlite3.OperationalError:
                pass
        # fired column (ignore error if already exists)
        try:
            conn.execute(
                "ALTER TABLE transactions ADD COLUMN fired INTEGER NOT NULL DEFAULT 0"
            )
        except sqlite3.OperationalError:
            pass
        # Producer columns may be missing on old simulation.db files.
        _migrate_simulation_db(conn)
        # Now create all indexes — columns are guaranteed to exist.
        for stmt in _TX_DDL.strip().split(";"):
            s = stmt.strip()
            if s and s.upper().startswith("CREATE INDEX"):
                try:
                    conn.execute(s)
                except sqlite3.OperationalError:
                    pass
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_fired "
            "ON transactions(fired, transaction_date)"
        )
    conn.close()


# ── European Custom DB write ───────────────────────────────────────────────────

_TX_INSERT_KEYS: list[str] = [
    "transaction_id", "transaction_date", "seller_id", "seller_name",
    "seller_country", "item_description", "item_category",
    "value", "vat_rate", "vat_amount", "buyer_country",
    "correct_vat_rate", "has_error", "xml_message", "created_at",
    "producer_id", "producer_name", "producer_country", "producer_city",
    # New-dataset (Stage 3) per-tx engine inputs/outputs. Optional —
    # NULL on legacy rows, populated by lib/new_seeder.py.
    "vat_subcategory_code",
    "engine_vat_ratio_risk",
    "engine_ml_risk",
    "engine_ml_seller_contribution",
    "engine_ml_origin_contribution",
    "engine_ml_category_contribution",
    "engine_ml_destination_contribution",
    "engine_vagueness_risk",
    "engine_ie_watchlist_risk",
]
_TX_INSERT_PLACEHOLDERS = ", ".join(f":{k}" for k in _TX_INSERT_KEYS)
_TX_INSERT_COLS         = ", ".join(_TX_INSERT_KEYS)
_TX_INSERT_NULLABLE_DEFAULTS = {k: None for k in _TX_INSERT_KEYS
                                if k not in ("transaction_id", "transaction_date",
                                             "seller_id", "seller_name", "seller_country",
                                             "item_description", "item_category",
                                             "value", "vat_rate", "vat_amount",
                                             "buyer_country", "correct_vat_rate",
                                             "has_error", "created_at")}


def _fill_tx_defaults(row: dict) -> dict:
    """Populate any missing nullable keys with None so the named-bind insert
    statements don't raise sqlite3.ProgrammingError on legacy callers."""
    for k, default in _TX_INSERT_NULLABLE_DEFAULTS.items():
        row.setdefault(k, default)
    return row


def insert_transaction(row: dict) -> None:
    _fill_tx_defaults(row)
    conn = _connect(EUROPEAN_CUSTOM_DB)
    with conn:
        conn.execute(
            f"""
            INSERT INTO transactions ({_TX_INSERT_COLS})
            VALUES ({_TX_INSERT_PLACEHOLDERS})
            ON CONFLICT(transaction_id) DO UPDATE SET
              transaction_date  = excluded.transaction_date,
              seller_name       = excluded.seller_name,
              item_description  = excluded.item_description,
              value             = excluded.value,
              vat_rate          = excluded.vat_rate,
              vat_amount        = excluded.vat_amount,
              correct_vat_rate  = excluded.correct_vat_rate,
              has_error         = excluded.has_error,
              producer_id       = excluded.producer_id,
              producer_name     = excluded.producer_name,
              producer_country  = excluded.producer_country,
              producer_city     = excluded.producer_city,
              vat_subcategory_code              = excluded.vat_subcategory_code,
              engine_vat_ratio_risk             = excluded.engine_vat_ratio_risk,
              engine_ml_risk                    = excluded.engine_ml_risk,
              engine_ml_seller_contribution     = excluded.engine_ml_seller_contribution,
              engine_ml_origin_contribution     = excluded.engine_ml_origin_contribution,
              engine_ml_category_contribution   = excluded.engine_ml_category_contribution,
              engine_ml_destination_contribution= excluded.engine_ml_destination_contribution,
              engine_vagueness_risk             = excluded.engine_vagueness_risk,
              engine_ie_watchlist_risk          = excluded.engine_ie_watchlist_risk
            """,
            row,
        )
    conn.close()


def bulk_insert(rows: list[dict], path: Path = EUROPEAN_CUSTOM_DB) -> None:
    rows = [_fill_tx_defaults(r) for r in rows]
    conn = _connect(path)
    with conn:
        conn.executemany(
            f"""
            INSERT OR IGNORE INTO transactions ({_TX_INSERT_COLS})
            VALUES ({_TX_INSERT_PLACEHOLDERS})
            """,
            rows,
        )
    conn.close()


# ── European Custom DB read ────────────────────────────────────────────────────

def get_latest_transactions(limit: int = 30) -> list[dict]:
    conn = _connect(EUROPEAN_CUSTOM_DB)
    rows = conn.execute(
        "SELECT * FROM transactions ORDER BY transaction_date DESC, id DESC LIMIT ?",
        (limit,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_transaction_by_id(transaction_id: str) -> dict | None:
    conn = _connect(EUROPEAN_CUSTOM_DB)
    row = conn.execute(
        "SELECT * FROM transactions WHERE transaction_id=? LIMIT 1",
        (transaction_id,),
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def get_transaction_count() -> int:
    conn = _connect(EUROPEAN_CUSTOM_DB)
    n = conn.execute("SELECT COUNT(*) FROM transactions").fetchone()[0]
    conn.close()
    return n


def query_transactions(
    *,
    seller_name: str | None = None,
    buyer_country: str | None = None,
    seller_country: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    limit: int = 500,
    offset: int = 0,
) -> list[dict]:
    clauses, params = [], []
    if seller_name:
        clauses.append("seller_name = ?")
        params.append(seller_name)
    if buyer_country:
        clauses.append("buyer_country = ?")
        params.append(buyer_country)
    if seller_country:
        clauses.append("seller_country = ?")
        params.append(seller_country)
    if date_from:
        clauses.append("transaction_date >= ?")
        params.append(date_from)
    if date_to:
        clauses.append("transaction_date <= ?")
        params.append(date_to + "T23:59:59")
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    sql = f"SELECT * FROM transactions {where} ORDER BY transaction_date DESC LIMIT ? OFFSET ?"
    params += [limit, offset]
    conn = _connect(EUROPEAN_CUSTOM_DB)
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_vat_metrics(
    *,
    seller_name: str | None = None,
    buyer_country: str | None = None,
    seller_country: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
) -> dict[str, Any]:
    """Aggregate VAT metrics with optional filters."""
    clauses, params = [], []
    if seller_name:
        clauses.append("seller_name = ?")
        params.append(seller_name)
    if buyer_country:
        clauses.append("buyer_country = ?")
        params.append(buyer_country)
    if seller_country:
        clauses.append("seller_country = ?")
        params.append(seller_country)
    if date_from:
        clauses.append("transaction_date >= ?")
        params.append(date_from)
    if date_to:
        clauses.append("transaction_date <= ?")
        params.append(date_to + "T23:59:59")
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    conn = _connect(EUROPEAN_CUSTOM_DB)

    totals = conn.execute(
        f"SELECT COUNT(*) as n, SUM(value) as total_value, "
        f"SUM(vat_amount) as total_vat, SUM(has_error) as errors "
        f"FROM transactions {where}",
        params,
    ).fetchone()

    by_buyer = conn.execute(
        f"SELECT buyer_country, COUNT(*) as n, SUM(vat_amount) as vat "
        f"FROM transactions {where} GROUP BY buyer_country ORDER BY vat DESC",
        params,
    ).fetchall()

    by_seller = conn.execute(
        f"SELECT seller_name, COUNT(*) as n, SUM(vat_amount) as vat "
        f"FROM transactions {where} GROUP BY seller_name ORDER BY vat DESC",
        params,
    ).fetchall()

    by_category = conn.execute(
        f"SELECT item_category, COUNT(*) as n, SUM(vat_amount) as vat "
        f"FROM transactions {where} GROUP BY item_category ORDER BY vat DESC",
        params,
    ).fetchall()

    # Daily VAT over time
    daily = conn.execute(
        f"SELECT SUBSTR(transaction_date,1,10) as day, SUM(vat_amount) as vat "
        f"FROM transactions {where} GROUP BY day ORDER BY day",
        params,
    ).fetchall()

    conn.close()
    return {
        "total_transactions": totals["n"] or 0,
        "total_value": round(totals["total_value"] or 0, 2),
        "total_vat": round(totals["total_vat"] or 0, 2),
        "error_count": totals["errors"] or 0,
        "by_buyer_country": [dict(r) for r in by_buyer],
        "by_seller": [dict(r) for r in by_seller],
        "by_category": [dict(r) for r in by_category],
        "daily_vat": [dict(r) for r in daily],
    }


# ── Simulation DB ─────────────────────────────────────────────────────────────

def get_next_sim_transaction() -> dict | None:
    """Return the single next unfired transaction in chronological order."""
    conn = _connect(SIMULATION_DB)
    row = conn.execute(
        "SELECT * FROM transactions WHERE fired=0 ORDER BY transaction_date LIMIT 1"
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def get_pending_sim_transactions(up_to_date: str, batch: int = 100) -> list[dict]:
    """Return unfired simulation transactions whose date <= up_to_date."""
    conn = _connect(SIMULATION_DB)
    rows = conn.execute(
        "SELECT * FROM transactions WHERE fired=0 AND transaction_date <= ? "
        "ORDER BY transaction_date LIMIT ?",
        (up_to_date, batch),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def mark_fired(transaction_ids: list[str]) -> None:
    conn = _connect(SIMULATION_DB)
    with conn:
        conn.executemany(
            "UPDATE transactions SET fired=1 WHERE transaction_id=?",
            [(tid,) for tid in transaction_ids],
        )
    conn.close()


def reset_simulation_db() -> None:
    conn = _connect(SIMULATION_DB)
    with conn:
        conn.execute("UPDATE transactions SET fired=0")
    conn.close()


def get_sim_counts() -> dict[str, int]:
    conn = _connect(SIMULATION_DB)
    total = conn.execute("SELECT COUNT(*) FROM transactions").fetchone()[0]
    fired = conn.execute("SELECT COUNT(*) FROM transactions WHERE fired=1").fetchone()[0]
    conn.close()
    return {"total": total, "fired": fired, "remaining": total - fired}


# ── Alarm queries (European Custom DB) ───────────────────────────────────────

def get_alarms(active_only: bool = False, limit: int = 50) -> list[dict]:
    conn = _connect(EUROPEAN_CUSTOM_DB)
    where = "WHERE active=1" if active_only else ""
    rows = conn.execute(
        f"SELECT * FROM alarms {where} ORDER BY id DESC LIMIT ?", (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def expire_old_alarms(as_of: str) -> None:
    """Deactivate alarms whose expiry has passed."""
    conn = _connect(EUROPEAN_CUSTOM_DB)
    with conn:
        conn.execute(
            "UPDATE alarms SET active=0 WHERE active=1 AND expires_at <= ?",
            (as_of,),
        )
    conn.close()


def get_suspicious_transactions(limit: int = 50) -> list[dict]:
    conn = _connect(EUROPEAN_CUSTOM_DB)
    rows = conn.execute(
        """
        SELECT t.*, a.deviation_pct, a.ratio_current, a.ratio_historical,
               a.raised_at as alarm_raised_at, a.expires_at as alarm_expires_at
        FROM transactions t
        JOIN alarms a ON t.alarm_id = a.id
        WHERE t.suspicious = 1
        ORDER BY t.transaction_date DESC
        LIMIT ?
        """,
        (limit,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def reset_alarms() -> None:
    """
    Prepare the European Custom DB for a fresh simulation run:
      - Remove simulation-period transactions (≥ 2026-03-01) so the pipeline
        can re-insert them with updated risk scores.
      - Clear alarms, agent log and ireland queue.
      - Reset suspicious flags on the retained historical records.
    Historical rows (Sep 2025 – Feb 2026) are kept intact as baseline context.
    """
    from lib.config import SIM_START_STR
    conn = _connect(EUROPEAN_CUSTOM_DB)
    with conn:
        conn.execute("DELETE FROM transactions WHERE transaction_date >= ?", (SIM_START_STR,))
        conn.execute("DELETE FROM alarms")
        conn.execute("DELETE FROM agent_log")
        conn.execute("DELETE FROM ireland_queue")
        conn.execute(
            "UPDATE transactions SET suspicious=0, alarm_id=NULL, suspicion_level=NULL"
        )
    conn.close()


def historical_transaction_count() -> int:
    """Number of pre-simulation transactions in the European Custom DB."""
    from lib.config import SIM_START_STR
    conn = _connect(EUROPEAN_CUSTOM_DB)
    n = conn.execute(
        "SELECT COUNT(*) FROM transactions WHERE transaction_date < ?", (SIM_START_STR,)
    ).fetchone()[0]
    conn.close()
    return n


# ── Agent log ─────────────────────────────────────────────────────────────────

def insert_agent_log(entry: dict) -> None:
    conn = _connect(EUROPEAN_CUSTOM_DB)
    try:
        with conn:
            conn.execute(
                """
                INSERT INTO agent_log
                (transaction_id, seller_name, buyer_country, item_description,
                 item_category, value, vat_rate, correct_vat_rate,
                 verdict, reasoning, legislation_refs, sent_to_ireland, processed_at)
                VALUES
                (:transaction_id, :seller_name, :buyer_country, :item_description,
                 :item_category, :value, :vat_rate, :correct_vat_rate,
                 :verdict, :reasoning, :legislation_refs, :sent_to_ireland, :processed_at)
                """,
                entry,
            )
    except Exception as e:
        print(f"  [agent_log] INSERT failed: {e}")
    finally:
        conn.close()


def get_agent_log(limit: int = 100) -> list[dict]:
    import json as _json
    conn = _connect(EUROPEAN_CUSTOM_DB)
    rows = conn.execute(
        "SELECT * FROM agent_log ORDER BY id DESC LIMIT ?", (limit,)
    ).fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        try:
            d["legislation_refs"] = _json.loads(d["legislation_refs"]) if d.get("legislation_refs") else []
        except Exception:
            d["legislation_refs"] = []
        result.append(d)
    return result


def get_agent_log_by_tx(transaction_id: str) -> dict | None:
    import json as _json
    conn = _connect(EUROPEAN_CUSTOM_DB)
    row = conn.execute(
        "SELECT * FROM agent_log WHERE transaction_id=? ORDER BY id DESC LIMIT 1",
        (transaction_id,),
    ).fetchone()
    conn.close()
    if not row:
        return None
    d = dict(row)
    try:
        d["legislation_refs"] = _json.loads(d["legislation_refs"]) if d.get("legislation_refs") else []
    except Exception:
        d["legislation_refs"] = []
    return d


def flag_transaction_suspicious(
    transaction_id: str,
    alarm_id: int | None,
    risk_score: str = "amber",
) -> None:
    """
    DB-subscriber action triggered by the Release_Event_Broker.
    Updates the stored transaction record using its identifier — sets
    suspicious=1, links the alarm (if any), and stores the computed
    risk_score ('amber' or 'red') as suspicion_level.
    """
    conn = _connect(EUROPEAN_CUSTOM_DB)
    with conn:
        conn.execute(
            "UPDATE transactions "
            "SET suspicious=1, alarm_id=?, suspicion_level=? "
            "WHERE transaction_id=?",
            (alarm_id, risk_score, transaction_id),
        )
    conn.close()


def update_suspicion_level(transaction_id: str, level: str) -> None:
    conn = _connect(EUROPEAN_CUSTOM_DB)
    with conn:
        conn.execute(
            "UPDATE transactions SET suspicion_level=? WHERE transaction_id=?",
            (level, transaction_id),
        )
    conn.close()


def clear_suspicious_flag(transaction_id: str) -> None:
    """Remove suspicious flag when agent clears the transaction."""
    conn = _connect(EUROPEAN_CUSTOM_DB)
    with conn:
        conn.execute(
            "UPDATE transactions SET suspicious=0, alarm_id=NULL, suspicion_level=NULL "
            "WHERE transaction_id=?",
            (transaction_id,),
        )
    conn.close()


# ── Ireland queue ─────────────────────────────────────────────────────────────

def insert_ireland_queue(entry: dict) -> None:
    conn = _connect(EUROPEAN_CUSTOM_DB)
    with conn:
        conn.execute(
            """
            INSERT OR IGNORE INTO ireland_queue
            (transaction_id, seller_name, seller_country, item_description,
             item_category, value, vat_rate, correct_vat_rate, vat_amount,
             transaction_date, alarm_key, deviation_pct, ratio_current,
             ratio_historical, agent_verdict, agent_reasoning, queued_at)
            VALUES
            (:transaction_id, :seller_name, :seller_country, :item_description,
             :item_category, :value, :vat_rate, :correct_vat_rate, :vat_amount,
             :transaction_date, :alarm_key, :deviation_pct, :ratio_current,
             :ratio_historical, :agent_verdict, :agent_reasoning, :queued_at)
            """,
            entry,
        )
    conn.close()


def get_ireland_queue(limit: int = 100) -> list[dict]:
    conn = _connect(EUROPEAN_CUSTOM_DB)
    rows = conn.execute(
        "SELECT * FROM ireland_queue ORDER BY id DESC LIMIT ?", (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_ireland_case(transaction_id: str) -> dict | None:
    """Return ireland_queue entry merged with agent_log detail (legislation refs)."""
    import json as _json
    conn = _connect(EUROPEAN_CUSTOM_DB)
    iq = conn.execute(
        "SELECT * FROM ireland_queue WHERE transaction_id=? LIMIT 1",
        (transaction_id,),
    ).fetchone()
    al = conn.execute(
        "SELECT * FROM agent_log WHERE transaction_id=? ORDER BY id DESC LIMIT 1",
        (transaction_id,),
    ).fetchone()
    conn.close()
    if not iq:
        return None
    result = dict(iq)
    if al:
        try:
            result["legislation_refs"] = _json.loads(al["legislation_refs"]) if al["legislation_refs"] else []
        except Exception:
            result["legislation_refs"] = []
        result["agent_reasoning_full"] = al["reasoning"]
    else:
        result["legislation_refs"] = []
    return result


# ── Sales_Order_Case read / update (investigation.db) ────────────────────────

def get_all_cases(status: str | None = None, limit: int = 200) -> list[dict]:
    """Return Sales_Order_Case rows, optionally filtered by Status."""
    import json as _json
    conn = _connect(INVESTIGATION_DB)
    if status:
        rows = conn.execute(
            "SELECT * FROM Sales_Order_Case WHERE Status = ? "
            "ORDER BY Update_time DESC LIMIT ?", (status, limit),
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM Sales_Order_Case ORDER BY Update_time DESC LIMIT ?",
            (limit,),
        ).fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        # Parse Communication JSON string into a list
        try:
            d["Communication"] = _json.loads(d["Communication"]) if d.get("Communication") else []
        except Exception:
            d["Communication"] = []
        result.append(d)
    return result


def get_case_by_id(case_id: str) -> dict | None:
    """Single case from investigation.db."""
    import json as _json
    conn = _connect(INVESTIGATION_DB)
    row = conn.execute(
        "SELECT * FROM Sales_Order_Case WHERE Case_ID = ? LIMIT 1",
        (case_id,),
    ).fetchone()
    conn.close()
    if not row:
        return None
    d = dict(row)
    try:
        d["Communication"] = _json.loads(d["Communication"]) if d.get("Communication") else []
    except Exception:
        d["Communication"] = []
    return d


def update_case(case_id: str, updates: dict) -> bool:
    """Partial update of a Sales_Order_Case row. Returns True if a row was updated."""
    import json as _json
    if not updates:
        return False
    # Serialize Communication back to JSON string if present
    if "Communication" in updates and isinstance(updates["Communication"], list):
        updates["Communication"] = _json.dumps(updates["Communication"])
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [case_id]
    conn = _connect(INVESTIGATION_DB)
    with conn:
        cur = conn.execute(
            f"UPDATE Sales_Order_Case SET {set_clause} WHERE Case_ID = ?",
            values,
        )
    changed = cur.rowcount > 0
    conn.close()
    return changed


# ── Reference table seed + getters ──────────────────────────────────────────

def _seed_reference_tables() -> None:
    """Idempotent seed of the four lookup tables. INSERT OR IGNORE keyed on
    the natural unique column so re-runs don't duplicate rows."""
    conn = _connect(EUROPEAN_CUSTOM_DB)
    with conn:
        conn.executemany(
            "INSERT OR IGNORE INTO vat_categories (label, rate, description, sort_order) VALUES (?, ?, ?, ?)",
            _SEED_VAT_CATEGORIES,
        )
        conn.executemany(
            "INSERT OR IGNORE INTO risk_levels (name, display_color, sort_order) VALUES (?, ?, ?)",
            _SEED_RISK_LEVELS,
        )
        conn.executemany(
            "INSERT OR IGNORE INTO eu_regions (country_code, country_name, region) VALUES (?, ?, ?)",
            _SEED_REGIONS,
        )
        conn.executemany(
            "INSERT OR IGNORE INTO suspicion_types (name, description, icon, color, sort_order) VALUES (?, ?, ?, ?, ?)",
            _SEED_SUSPICION_TYPES,
        )
        conn.executemany(
            "INSERT OR IGNORE INTO sales_order_statuses (name, description, sort_order) VALUES (?, ?, ?)",
            _SEED_SALES_ORDER_STATUSES,
        )
        conn.executemany(
            "INSERT OR IGNORE INTO case_statuses (name, description, sort_order) VALUES (?, ?, ?)",
            _SEED_CASE_STATUSES,
        )
    conn.close()


def get_vat_categories() -> list[dict]:
    conn = _connect(EUROPEAN_CUSTOM_DB)
    rows = conn.execute(
        "SELECT label, rate, description FROM vat_categories ORDER BY sort_order, label"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_risk_levels() -> list[dict]:
    conn = _connect(EUROPEAN_CUSTOM_DB)
    rows = conn.execute(
        "SELECT name, display_color FROM risk_levels ORDER BY sort_order, name"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_eu_regions() -> list[dict]:
    conn = _connect(EUROPEAN_CUSTOM_DB)
    rows = conn.execute(
        "SELECT country_code, country_name, region FROM eu_regions ORDER BY region, country_code"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ── ML risk-rules table (4-tuple lookup + per-dimension weights) ────────────
#
# Single mapping table driving two things:
#   1. Risk monitoring engine 2 (watchlist): looks up the 4-tuple
#      (seller, country_origin, vat_product_category, country_destination)
#      and uses `risk` to decide whether to flag.
#   2. Case creation: populates Sales_Order_Risk dimensional scores from
#      the four weight columns.
# Source of truth: context/Fake ML.xlsx. Re-seeded on every backend start
# so edits to the spreadsheet propagate after a restart.

_ML_XLSX_PATH = Path(__file__).parent.parent / "context" / "Fake ML.xlsx"


def _seed_ml_risk_rules_from_xlsx() -> None:
    """Clear the ml_risk_rules table and re-load it from Fake ML.xlsx.

    Silent no-op if the spreadsheet (or openpyxl) is missing — the engine
    then simply flags nothing.
    """
    if not _ML_XLSX_PATH.exists():
        return
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return

    wb = openpyxl.load_workbook(_ML_XLSX_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return
    # row 0 = header; subsequent rows are data.
    conn = _connect(EUROPEAN_CUSTOM_DB)
    with conn:
        conn.execute("DELETE FROM ml_risk_rules")
        for r in rows[1:]:
            # Skip blank rows (openpyxl gives a tuple of Nones for empty ones)
            if r is None or all(c is None for c in r):
                continue
            r = list(r) + [None] * max(0, 10 - len(r))
            conn.execute("""
                INSERT OR REPLACE INTO ml_risk_rules (
                    seller, country_origin, vat_product_category, country_destination,
                    risk, description,
                    seller_weight, country_origin_weight,
                    vat_product_category_weight, country_destination_weight
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                (r[0] or "").strip(), (r[1] or "").strip(),
                (r[2] or "").strip(), (r[3] or "").strip(),
                float(r[4]) if r[4] is not None else 0.0,
                r[5],
                float(r[6]) if r[6] is not None else None,
                float(r[7]) if r[7] is not None else None,
                float(r[8]) if r[8] is not None else None,
                float(r[9]) if r[9] is not None else None,
            ))
    conn.close()


def lookup_ml_risk_rule(seller: str, country_origin: str,
                        vat_product_category: str, country_destination: str) -> dict | None:
    """Case-insensitive 4-tuple lookup. Returns the rule dict or None."""
    conn = _connect(EUROPEAN_CUSTOM_DB)
    row = conn.execute("""
        SELECT * FROM ml_risk_rules
        WHERE LOWER(seller)               = LOWER(?)
          AND LOWER(country_origin)       = LOWER(?)
          AND LOWER(vat_product_category) = LOWER(?)
          AND LOWER(country_destination)  = LOWER(?)
        LIMIT 1
    """, (seller or "", country_origin or "",
          vat_product_category or "", country_destination or "")).fetchone()
    conn.close()
    return dict(row) if row else None


def get_case_statuses() -> list[dict]:
    conn = _connect(EUROPEAN_CUSTOM_DB)
    rows = conn.execute(
        "SELECT name, description FROM case_statuses ORDER BY sort_order, name"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_sales_order_statuses() -> list[dict]:
    conn = _connect(EUROPEAN_CUSTOM_DB)
    rows = conn.execute(
        "SELECT name, description FROM sales_order_statuses ORDER BY sort_order, name"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_suspicion_types() -> list[dict]:
    conn = _connect(EUROPEAN_CUSTOM_DB)
    rows = conn.execute(
        "SELECT name, description, icon, color FROM suspicion_types ORDER BY sort_order, name"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ── Case grouping: similarity check + append ────────────────────────────────
#
# Similar transactions (exact seller + destination + category, fuzzy
# description) are grouped into the same open case. Similarity is
# assessed with Jaccard word overlap on the product description.

DESCRIPTION_SIMILARITY_THRESHOLD = 0.4


def _jaccard_words(a: str, b: str) -> float:
    """Jaccard similarity on lowercased word sets."""
    wa = set((a or "").lower().split())
    wb = set((b or "").lower().split())
    if not wa or not wb:
        return 0.0
    return len(wa & wb) / len(wa | wb)


def get_previous_cases(seller: str, exclude_case_id: str = "",
                       limit: int = 20) -> list[dict]:
    """Return past CLOSED cases from the same seller.

    Reads from historical_cases.db (populated once by
    lib.historical_seeder). That DB is independent of investigation.db
    — the point of "previous cases" is to expose curated past
    investigations, not just the cases the current sim happens to
    have closed."""
    conn = _connect(HISTORICAL_CASES_DB)
    rows = conn.execute("""
        SELECT c.Case_ID, c.Status, c.VAT_Problem_Type,
               c.Overall_Case_Risk_Score, c.Overall_Case_Risk_Level,
               c.Created_time, c.Update_time,
               o.Seller_Name, o.Country_Origin, o.Country_Destination,
               o.HS_Product_Category, o.Product_Description,
               (SELECT COUNT(*) FROM Sales_Order s2 WHERE s2.Case_ID = c.Case_ID) AS order_count,
               c.Proposed_Action_Customs,
               c.Proposed_Action_Tax
        FROM Sales_Order_Case c
        LEFT JOIN Sales_Order o ON c.Sales_Order_Business_Key = o.Sales_Order_Business_Key
        WHERE o.Seller_Name = ? AND c.Case_ID != ? AND c.Status = 'Closed'
        ORDER BY c.Update_time DESC LIMIT ?
    """, (seller, exclude_case_id, limit)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_correlated_cases(seller: str, category: str, destination: str,
                         exclude_case_id: str = "", limit: int = 20) -> list[dict]:
    """Return OPEN cases matching (seller, declared category, destination).

    Tightened correlation key per slide 1 of Rules in App.pptx: a case
    is "correlated" with the current one if it shares the same seller,
    same declared product category, AND same destination — and is still
    under investigation (not closed). Previously keyed on category only."""
    conn = _connect(INVESTIGATION_DB)
    rows = conn.execute("""
        SELECT c.Case_ID, c.Status, c.VAT_Problem_Type,
               c.Overall_Case_Risk_Score, c.Overall_Case_Risk_Level,
               c.Created_time,
               o.Seller_Name, o.Country_Origin, o.Country_Destination,
               o.HS_Product_Category, o.Product_Description,
               (SELECT COUNT(*) FROM Sales_Order s2 WHERE s2.Case_ID = c.Case_ID) AS order_count
        FROM Sales_Order_Case c
        LEFT JOIN Sales_Order o ON c.Sales_Order_Business_Key = o.Sales_Order_Business_Key
        WHERE o.Seller_Name         = ?
          AND o.HS_Product_Category = ?
          AND o.Country_Destination = ?
          AND c.Case_ID != ? AND c.Status != 'Closed'
        ORDER BY c.Overall_Case_Risk_Score DESC LIMIT ?
    """, (seller, category, destination, exclude_case_id, limit)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def find_similar_open_case(
    seller: str, destination: str, category: str, description: str,
) -> dict | None:
    """Find an open (non-Closed) case whose primary or grouped orders
    share the same seller + destination + category AND whose product
    description passes the Jaccard similarity threshold.

    Returns the hydrated case dict or None.
    """
    conn = _connect(INVESTIGATION_DB)
    # Query all Sales_Orders linked to non-Closed cases that match
    # the exact fields. Case_ID on Sales_Order is set at creation.
    rows = conn.execute("""
        SELECT DISTINCT o.Case_ID, o.Product_Description, c.Status
        FROM Sales_Order o
        JOIN Sales_Order_Case c ON o.Case_ID = c.Case_ID
        WHERE o.Seller_Name          = ?
          AND o.Country_Destination  = ?
          AND o.HS_Product_Category  = ?
          AND c.Status               != 'Closed'
          AND o.Case_ID IS NOT NULL
    """, (seller or "", destination or "", category or "")).fetchall()
    conn.close()

    best_case_id = None
    best_sim     = 0.0
    for r in rows:
        sim = _jaccard_words(description, r["Product_Description"])
        if sim >= DESCRIPTION_SIMILARITY_THRESHOLD and sim > best_sim:
            best_sim     = sim
            best_case_id = r["Case_ID"]

    if best_case_id is None:
        return None
    return get_case_hydrated(best_case_id)


def append_order_to_case(case_id: str, so_row: dict, sor_row: dict) -> None:
    """Add a Sales_Order + Sales_Order_Risk to an existing case.
    Sets Case_ID on the Sales_Order row. Single transaction."""
    so_row["Case_ID"] = case_id
    conn = _connect(INVESTIGATION_DB)
    try:
        with conn:
            conn.execute("""
                INSERT OR REPLACE INTO Sales_Order (
                    Sales_Order_ID, Sales_Order_Business_Key,
                    HS_Product_Category, Product_Description, Product_Value,
                    VAT_Rate, VAT_Fee, Seller_Name,
                    Country_Origin, Country_Destination,
                    Status, Update_time, Updated_by, Case_ID
                ) VALUES (
                    :Sales_Order_ID, :Sales_Order_Business_Key,
                    :HS_Product_Category, :Product_Description, :Product_Value,
                    :VAT_Rate, :VAT_Fee, :Seller_Name,
                    :Country_Origin, :Country_Destination,
                    :Status, :Update_time, :Updated_by, :Case_ID
                )
            """, so_row)
            conn.execute("""
                INSERT OR REPLACE INTO Sales_Order_Risk (
                    Sales_Order_Risk_ID, Sales_Order_Business_Key,
                    Risk_Type, Overall_Risk_Score, Overall_Risk_Level,
                    Seller_Risk_Score, Country_Risk_Score,
                    Product_Category_Risk_Score, Manufacturer_Risk_Score,
                    Confidence_Score, Overall_Risk_Description,
                    Proposed_Risk_Action, Risk_Comment,
                    Evaluation_by, Update_time, Updated_by
                ) VALUES (
                    :Sales_Order_Risk_ID, :Sales_Order_Business_Key,
                    :Risk_Type, :Overall_Risk_Score, :Overall_Risk_Level,
                    :Seller_Risk_Score, :Country_Risk_Score,
                    :Product_Category_Risk_Score, :Manufacturer_Risk_Score,
                    :Confidence_Score, :Overall_Risk_Description,
                    :Proposed_Risk_Action, :Risk_Comment,
                    :Evaluation_by, :Update_time, :Updated_by
                )
            """, sor_row)
    finally:
        conn.close()


def update_case_engine_scores(case_id: str, engine_scores: dict,
                              overall_score: float, risk_level: str) -> None:
    """Update the per-engine and overall risk scores on a case."""
    conn = _connect(INVESTIGATION_DB)
    with conn:
        conn.execute("""
            UPDATE Sales_Order_Case SET
                Overall_Case_Risk_Score       = :Overall_Case_Risk_Score,
                Overall_Case_Risk_Level       = :Overall_Case_Risk_Level,
                Engine_VAT_Ratio              = :Engine_VAT_Ratio,
                Engine_ML_Watchlist            = :Engine_ML_Watchlist,
                Engine_IE_Seller_Watchlist     = :Engine_IE_Seller_Watchlist,
                Engine_Description_Vagueness   = :Engine_Description_Vagueness,
                Update_time                    = :Update_time
            WHERE Case_ID = :Case_ID
        """, {**engine_scores,
              "Overall_Case_Risk_Score": overall_score,
              "Overall_Case_Risk_Level": risk_level,
              "Case_ID": case_id,
              "Update_time": datetime.now(timezone.utc).isoformat()})
    conn.close()


def get_risk_engine_signals() -> list[dict]:
    """Return the reference table mapping field names to display names."""
    conn = _connect(INVESTIGATION_DB)
    rows = conn.execute("SELECT * FROM risk_engine_signals").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_case_transaction_count(case_id: str) -> int:
    """Count Sales_Order rows linked to a case."""
    conn = _connect(INVESTIGATION_DB)
    n = conn.execute(
        "SELECT COUNT(*) FROM Sales_Order WHERE Case_ID = ?", (case_id,)
    ).fetchone()[0]
    conn.close()
    return n


def update_sales_order_status(business_key: str, status: str) -> bool:
    """Update Sales_Order.Status in investigation.db for a given business key."""
    conn = _connect(INVESTIGATION_DB)
    with conn:
        cur = conn.execute(
            "UPDATE Sales_Order SET Status = ?, Update_time = ? WHERE Sales_Order_Business_Key = ?",
            (status, datetime.now(timezone.utc).isoformat(), business_key),
        )
    changed = cur.rowcount > 0
    conn.close()
    return changed


def seed_open_cases_if_empty() -> int:
    """Copy all rows from data/seed_cases.db into investigation.db when
    the latter currently has zero cases. Idempotent: a no-op if cases
    already exist or if the seed file is missing.

    Returns the number of cases inserted (0 if no seeding occurred).
    """
    if not SEED_CASES_DB.exists():
        return 0
    target = _connect(INVESTIGATION_DB)
    try:
        n = target.execute("SELECT COUNT(*) FROM Sales_Order_Case").fetchone()[0]
        if n > 0:
            return 0
        src = sqlite3.connect(SEED_CASES_DB)
        src.row_factory = sqlite3.Row
        try:
            with target:
                # Order matters because of the FK from Risk → Order
                for table in ("Sales_Order", "Sales_Order_Risk", "Sales_Order_Case"):
                    rows = src.execute(f"SELECT * FROM {table}").fetchall()
                    for r in rows:
                        cols = ",".join(r.keys())
                        placeholders = ",".join("?" * len(r.keys()))
                        target.execute(
                            f"INSERT OR REPLACE INTO {table} ({cols}) VALUES ({placeholders})",
                            list(r),
                        )
            inserted = target.execute("SELECT COUNT(*) FROM Sales_Order_Case").fetchone()[0]
            return inserted
        finally:
            src.close()
    finally:
        target.close()


def reset_cases() -> None:
    """Clear all 3 case-side tables in investigation.db (simulation reset)."""
    conn = _connect(INVESTIGATION_DB)
    with conn:
        conn.execute("DELETE FROM Sales_Order_Case")
        conn.execute("DELETE FROM Sales_Order_Risk")
        conn.execute("DELETE FROM Sales_Order")
    conn.close()


# ── Atomic 3-row insert into investigation.db ────────────────────────────────
#
# Used by the C&T Risk Management Factory when a retain/investigate
# ASSESSMENT_OUTCOME arrives. All three rows succeed or none do.

def upsert_investigation_set(so_row: dict, sor_row: dict, soc_row: dict) -> None:
    """Atomic insert of Sales_Order + Sales_Order_Risk + Sales_Order_Case
    into investigation.db. Single transaction."""
    conn = _connect(INVESTIGATION_DB)
    try:
        with conn:
            conn.execute("""
                INSERT OR REPLACE INTO Sales_Order (
                    Sales_Order_ID, Sales_Order_Business_Key,
                    HS_Product_Category, Product_Description, Product_Value,
                    VAT_Rate, VAT_Fee, Seller_Name,
                    Country_Origin, Country_Destination,
                    Status, Update_time, Updated_by, Case_ID
                ) VALUES (
                    :Sales_Order_ID, :Sales_Order_Business_Key,
                    :HS_Product_Category, :Product_Description, :Product_Value,
                    :VAT_Rate, :VAT_Fee, :Seller_Name,
                    :Country_Origin, :Country_Destination,
                    :Status, :Update_time, :Updated_by, :Case_ID
                )
            """, so_row)
            conn.execute("""
                INSERT OR REPLACE INTO Sales_Order_Risk (
                    Sales_Order_Risk_ID, Sales_Order_Business_Key,
                    Risk_Type, Overall_Risk_Score, Overall_Risk_Level,
                    Seller_Risk_Score, Country_Risk_Score,
                    Product_Category_Risk_Score, Manufacturer_Risk_Score,
                    Confidence_Score, Overall_Risk_Description,
                    Proposed_Risk_Action, Risk_Comment,
                    Evaluation_by, Update_time, Updated_by
                ) VALUES (
                    :Sales_Order_Risk_ID, :Sales_Order_Business_Key,
                    :Risk_Type, :Overall_Risk_Score, :Overall_Risk_Level,
                    :Seller_Risk_Score, :Country_Risk_Score,
                    :Product_Category_Risk_Score, :Manufacturer_Risk_Score,
                    :Confidence_Score, :Overall_Risk_Description,
                    :Proposed_Risk_Action, :Risk_Comment,
                    :Evaluation_by, :Update_time, :Updated_by
                )
            """, sor_row)
            conn.execute("""
                INSERT OR REPLACE INTO Sales_Order_Case (
                    Case_ID, Sales_Order_Business_Key, Status,
                    VAT_Problem_Type, Recommended_Product_Value,
                    Recommended_VAT_Product_Category, Recommended_VAT_Rate,
                    Recommended_VAT_Fee, AI_Analysis, AI_Confidence,
                    VAT_Gap_Fee, Evaluation_by,
                    Proposed_Action_Tax, Proposed_Action_Customs,
                    Communication, Additional_Evidence,
                    Update_time, Updated_by, Created_time,
                    Overall_Case_Risk_Score, Overall_Case_Risk_Level,
                    Engine_VAT_Ratio, Engine_ML_Watchlist,
                    Engine_IE_Seller_Watchlist, Engine_Description_Vagueness
                ) VALUES (
                    :Case_ID, :Sales_Order_Business_Key, :Status,
                    :VAT_Problem_Type, :Recommended_Product_Value,
                    :Recommended_VAT_Product_Category, :Recommended_VAT_Rate,
                    :Recommended_VAT_Fee, :AI_Analysis, :AI_Confidence,
                    :VAT_Gap_Fee, :Evaluation_by,
                    :Proposed_Action_Tax, :Proposed_Action_Customs,
                    :Communication, :Additional_Evidence,
                    :Update_time, :Updated_by, :Created_time,
                    :Overall_Case_Risk_Score, :Overall_Case_Risk_Level,
                    :Engine_VAT_Ratio, :Engine_ML_Watchlist,
                    :Engine_IE_Seller_Watchlist, :Engine_Description_Vagueness
                )
            """, soc_row)
    finally:
        conn.close()


_HYDRATED_CASE_SQL = """
    SELECT
        c.*,
        o.Sales_Order_ID,
        o.HS_Product_Category, o.Product_Description, o.Product_Value,
        o.VAT_Rate, o.VAT_Fee, o.Seller_Name,
        o.Country_Origin, o.Country_Destination,
        r.Sales_Order_Risk_ID,
        r.Risk_Type,
        r.Overall_Risk_Score, r.Overall_Risk_Level,
        r.Seller_Risk_Score, r.Country_Risk_Score,
        r.Product_Category_Risk_Score, r.Manufacturer_Risk_Score,
        r.Confidence_Score, r.Proposed_Risk_Action,
        r.Overall_Risk_Description,
        (SELECT COUNT(*) FROM Sales_Order s2 WHERE s2.Case_ID = c.Case_ID) AS transaction_count
    FROM Sales_Order_Case c
    LEFT JOIN Sales_Order      o ON c.Sales_Order_Business_Key = o.Sales_Order_Business_Key
    LEFT JOIN Sales_Order_Risk r ON c.Sales_Order_Business_Key = r.Sales_Order_Business_Key
"""


def _hydrate_row(r) -> dict:
    import json as _json
    d = dict(r)
    try:
        d["Communication"] = _json.loads(d["Communication"]) if d.get("Communication") else []
    except Exception:
        d["Communication"] = []
    return d


def get_case_orders(case_id: str) -> list[dict]:
    """Return all Sales_Order + Sales_Order_Risk rows for a case."""
    conn = _connect(INVESTIGATION_DB)
    rows = conn.execute("""
        SELECT o.Sales_Order_ID, o.Sales_Order_Business_Key,
               o.HS_Product_Category, o.Product_Description, o.Product_Value,
               o.VAT_Rate, o.VAT_Fee, o.Seller_Name,
               o.Country_Origin, o.Country_Destination,
               r.Overall_Risk_Score, r.Overall_Risk_Level,
               r.Seller_Risk_Score, r.Country_Risk_Score,
               r.Product_Category_Risk_Score, r.Manufacturer_Risk_Score,
               r.Confidence_Score
        FROM Sales_Order o
        LEFT JOIN Sales_Order_Risk r ON o.Sales_Order_Business_Key = r.Sales_Order_Business_Key
        WHERE o.Case_ID = ?
        ORDER BY o.Sales_Order_ID
    """, (case_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def _hydrate_with_orders(r, conn) -> dict:
    """Hydrate a case row and attach all its orders."""
    d = _hydrate_row(r)
    case_id = d.get("Case_ID")
    if case_id:
        orders = conn.execute("""
            SELECT o.Sales_Order_ID, o.Sales_Order_Business_Key,
                   o.HS_Product_Category, o.Product_Description, o.Product_Value,
                   o.VAT_Rate, o.VAT_Fee, o.Seller_Name,
                   o.Country_Origin, o.Country_Destination,
                   r.Overall_Risk_Score, r.Overall_Risk_Level
            FROM Sales_Order o
            LEFT JOIN Sales_Order_Risk r ON o.Sales_Order_Business_Key = r.Sales_Order_Business_Key
            WHERE o.Case_ID = ?
            ORDER BY o.Sales_Order_ID
        """, (case_id,)).fetchall()
        d["orders"] = [dict(o) for o in orders]
    else:
        d["orders"] = []
    return d


def get_all_cases_hydrated(status: str | None = None, limit: int = 200) -> list[dict]:
    """Return cases joined with Sales_Order + Sales_Order_Risk, with all orders attached."""
    conn = _connect(INVESTIGATION_DB)
    if status:
        sql = _HYDRATED_CASE_SQL + " WHERE c.Status = ? ORDER BY c.Created_time ASC LIMIT ?"
        rows = conn.execute(sql, (status, limit)).fetchall()
    else:
        sql = _HYDRATED_CASE_SQL + " ORDER BY c.Created_time ASC LIMIT ?"
        rows = conn.execute(sql, (limit,)).fetchall()
    result = [_hydrate_with_orders(r, conn) for r in rows]
    conn.close()
    return result


def get_case_hydrated(case_id: str) -> dict | None:
    """Single case joined with Sales_Order + Sales_Order_Risk, with all orders attached."""
    conn = _connect(INVESTIGATION_DB)
    sql = _HYDRATED_CASE_SQL + " WHERE c.Case_ID = ? LIMIT 1"
    row = conn.execute(sql, (case_id,)).fetchone()
    if not row:
        conn.close()
        return None
    result = _hydrate_with_orders(row, conn)
    conn.close()
    return result


# ── Data hub upserts (3 dark-purple tables) ──────────────────────────────────

